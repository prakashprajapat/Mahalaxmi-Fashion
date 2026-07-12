#!/usr/bin/env python3
"""
Payment Reconciliation → Excel
Razorpay payments vs site orders match karke xlsx banata hai.

VPS pe chalao:
    pip install requests openpyxl
    export MFH_ADMIN_EMAIL='admin@mahalaxmifashionhub.com'
    export MFH_ADMIN_PASSWORD='...'
    python3 reconcile_payments.py --from 2026-07-01 --to 2026-07-13

Cron (har mahine ki 1 tarikh, pichhla mahina):
    0 6 1 * * cd /var/www/mahalaxmi-nextjs && python3 reconcile_payments.py --last-month
"""
import argparse
import os
import sys
from datetime import date, timedelta

import requests

API = os.environ.get("MFH_API_BASE", "http://localhost:5000/api")


def get_args():
    p = argparse.ArgumentParser()
    p.add_argument("--from", dest="frm", help="YYYY-MM-DD")
    p.add_argument("--to", dest="to", help="YYYY-MM-DD")
    p.add_argument("--last-month", action="store_true")
    a = p.parse_args()
    if a.last_month:
        first_this = date.today().replace(day=1)
        last_prev = first_this - timedelta(days=1)
        return last_prev.replace(day=1).isoformat(), last_prev.isoformat()
    frm = a.frm or (date.today() - timedelta(days=30)).isoformat()
    to = a.to or date.today().isoformat()
    return frm, to


def admin_token():
    email = os.environ.get("MFH_ADMIN_EMAIL")
    password = os.environ.get("MFH_ADMIN_PASSWORD")
    if not email or not password:
        sys.exit("MFH_ADMIN_EMAIL / MFH_ADMIN_PASSWORD env vars set karo")
    r = requests.post(f"{API}/auth/admin-login", json={"email": email, "password": password}, timeout=30)
    r.raise_for_status()
    return r.json()["token"]


def main():
    frm, to = get_args()
    print(f"Reconciling {frm} → {to} ...")
    token = admin_token()
    r = requests.get(
        f"{API}/payments/reconcile",
        params={"from": frm, "to": to},
        headers={"Authorization": f"Bearer {token}"},
        timeout=300,
    )
    r.raise_for_status()
    data = r.json()
    s, rows = data["summary"], data["rows"]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    for k, label in [
        ("totalPayments", "Total Payments"), ("matched", "Matched"),
        ("amountMismatch", "Amount Mismatch"), ("paymentWithoutOrder", "Payment without Order"),
        ("orderWithoutPayment", "Order without Payment"), ("refunded", "Refunded"),
        ("failed", "Failed Payments"), ("capturedTotal", "Captured Total (Rs)"),
        ("refundedTotal", "Refunded Total (Rs)"),
    ]:
        ws.append([label, s[k]])
    for c in ws["A"] + ws[1]:
        c.font = Font(bold=True)

    # Detail sheet
    ws2 = wb.create_sheet("Reconciliation")
    headers = ["Category", "Payment ID", "Payment Rs", "Refunded Rs", "Payment Status",
               "Payment Date", "Order ID", "Order Rs", "Order Status", "Email", "Phone"]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    red = PatternFill("solid", fgColor="FFCDD2")
    yellow = PatternFill("solid", fgColor="FFF9C4")
    for row in rows:
        ws2.append([
            row["category"], row.get("paymentId"), row.get("paymentAmount"),
            row.get("refundedAmount"), row.get("paymentStatus"),
            (row.get("paymentDate") or "")[:19].replace("T", " "),
            row.get("orderId"), row.get("orderTotal"), row.get("orderStatus"),
            row.get("email"), row.get("contact"),
        ])
        fill = red if row["category"] in ("PAYMENT_NO_ORDER", "ORDER_NO_PAYMENT") \
            else yellow if row["category"] == "AMOUNT_MISMATCH" else None
        if fill:
            for c in ws2[ws2.max_row]:
                c.fill = fill

    for col, w in zip("ABCDEFGHIJK", [20, 22, 12, 12, 14, 20, 20, 10, 16, 26, 15]):
        ws2.column_dimensions[col].width = w

    out = f"payment-reconcile-{frm}-to-{to}.xlsx"
    wb.save(out)
    print(f"✅ {out}")
    print(f"   Matched: {s['matched']} | Mismatch: {s['amountMismatch']} | "
          f"Pay-no-order: {s['paymentWithoutOrder']} | Order-no-pay: {s['orderWithoutPayment']}")
    if s["amountMismatch"] or s["paymentWithoutOrder"] or s["orderWithoutPayment"]:
        print("⚠️  Mismatch rows hain — Excel me red/yellow rows check karo!")


if __name__ == "__main__":
    main()
